import openpyxl
from openpyxl.styles import Font

def add_ID_tab(wbIn):
    if "Example0gross_NameIDRecon" not in wbIn.worksheets:
        # only create the sheet if it is not present
        wbIn.create_sheet("Example0gross_NameIDRecon")
    ws_phdwin_id = wbIn.worksheets[1] # the sheet that will contain ID's
    # clear the sheet if it is not newly created before copying contents
    for row in ws_phdwin_id.iter_rows():
        for cell in row:
            cell.value = None

    # to copy the contents into the main workbook tab, open the source from open refine outputs
    wb_phdwin_id_source = openpyxl.load_workbook("./example_los/OpenRefine Outputs/Example0gross_NameIDReconciliation.xlsx")
    ws_phdwin_id_source = wb_phdwin_id_source.active # source from which to be copied

    for row in ws_phdwin_id_source.iter_rows(values_only=True):
        ws_phdwin_id.append(row)

    # fix top row for id tab
    ws_phdwin_id.freeze_panes = 'A2'

    # bold the headings
    ws_phdwin_id['A1'].font = ws_phdwin_id['B1'].font = ws_phdwin_id['C1'].font = Font(bold=True)

    wb_phdwin_id_source.close() # close the source workbook as it is no longer needed


def step_1():
    wbIn = openpyxl.load_workbook("./bot_outputs/step_0_out.xlsx") # the main workbook
    add_ID_tab(wbIn)
    
    wsIn = wbIn.worksheets[0] # the main sheet on which the process will be done

    wsIn.insert_cols(1)
    wsIn['A1'].value = 'PHDWIN Id'
    wsIn['A1'].font = Font(bold=True)

    # now to populate the phdwin ids
    for i in range(2, wsIn.max_row + 1):
        # formula
        wsIn.cell(row=i, column=1).value = f'=VLOOKUP($B{str(i)},Example0gross_NameIDRecon!$B:$C,2,0)'

    wbIn.save("./bot_outputs/step_1_out.xlsx")

step_1()