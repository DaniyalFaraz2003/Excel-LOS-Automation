import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule

def add_los_designation_tab(wbIn):
    if "Example0gross_LOSDesignation" not in wbIn.worksheets:
        # only create the sheet if it is not present
        wbIn.create_sheet("Example0gross_LOSDesignation")
    ws_los_designation = wbIn.worksheets[2] # the sheet that will contain Designations
    # clear the sheet if it is not newly created before copying contents
    for row in ws_los_designation.iter_rows():
        for cell in row:
            cell.value = None

    # to copy the contents into the main workbook tab, open the source from open refine outputs
    wb_los_designation_source = openpyxl.load_workbook("./example_los/OpenRefine Outputs/Example0gross_LOSDesignation.xlsx")
    ws_los_designation_source = wb_los_designation_source.active # source from which to be copied

    for row in ws_los_designation_source.iter_rows(values_only=True):
        ws_los_designation.append(row)

    # fix the top row and bold the headings and remove the text in C column (as shown in example)
    ws_los_designation['C1'].value = None
    ws_los_designation.freeze_panes = 'A2'
    ws_los_designation['A1'].font = ws_los_designation['B1'].font = ws_los_designation['D1'].font = ws_los_designation['D5'].font = Font(bold=True)

    # add formula for total
    ws_los_designation['E5'].value = '=SUM(E2:E4)'

    # edit number format to percentage
    for row in ws_los_designation['E2':'E5']:
        for cell in row:
            cell.number_format = '0%'

    # conditional formatting for sum
    white_font = Font(color='FFFFFF', bold=True) # color to be bold and white when red
    redFill = PatternFill(start_color='EE1111', end_color='EE1111', fill_type='solid')
    formula_rule = FormulaRule(formula=['=OR(AND(SUM(E2:E4)<>1, SUM(E2:E4)>0))'], stopIfTrue=False, fill=redFill, font=white_font)
    ws_los_designation.conditional_formatting.add('E5', formula_rule)

    wb_los_designation_source.close() # close the source workbook as it is no longer needed

def step_2():
    wbIn = openpyxl.load_workbook("./bot_outputs/step_1_out.xlsx") # main workbook
    add_los_designation_tab(wbIn)

    wsIn = wbIn.worksheets[0] # the main sheet on which the process will be done

    wsIn.insert_cols(4) # insert column at 4 for designations

    wsIn['D1'].value = 'LOS Designation'
    wsIn['D1'].font = Font(bold=True)

    # =IF(VLOOKUP($C2,Example0gross_LOSDesignation!$A:$B,2,0)=0,"",VLOOKUP($C2,Example0gross_LOSDesignation!$A:$B,2,0))
    # to populate Designations
    for i in range(2, wsIn.max_row + 1):
        # formula
        wsIn.cell(row=i, column=4).value = f'=IF(VLOOKUP($C{str(i)},Example0gross_LOSDesignation!$A:$B,2,0)=0,"",VLOOKUP($C{str(i)},Example0gross_LOSDesignation!$A:$B,2,0))'

    wbIn.save("./bot_outputs/step_2_out.xlsx")

step_2()