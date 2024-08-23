import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.utils import column_index_from_string
from datetime import datetime
from step_0 import copy_cell_styles, get_number_line_items

def add_btu_tab(wbIn):
    if "BTU" not in wbIn.worksheets:
        # only create the sheet if it is not present
        wbIn.create_sheet("BTU")
    ws_btu = wbIn.worksheets[4]  # the sheet that will contain BTU
    # clear the sheet if it is not newly created before copying contents
    for row in ws_btu.iter_rows():
        for cell in row:
            cell.value = None

    # to copy the contents into the main workbook tab, open the source from open refine outputs
    wb_btu_source = openpyxl.load_workbook("./OpenRefine Outputs/BTU.xlsx")
    ws_btu_source = wb_btu_source.active  # source from which to be copied

    for row in ws_btu_source.iter_rows(values_only=True):
        ws_btu.append(row)

    # correct formatting
    for i in range(1, ws_btu.max_row + 1):
        for j in range(1, ws_btu.max_column + 1):
            copy_cell_styles(ws_btu_source.cell(i, j), ws_btu.cell(i, j))

    # fix top row for btu tab
    ws_btu.freeze_panes = 'A2'

    # bold the headings
    ws_btu['A1'].font = ws_btu['B1'].font = ws_btu['C1'].font = Font(bold=True)

    wb_btu_source.close()  # close the source workbook as it is no longer needed

def add_historical_prod_tab(wbIn):
    if "Historical_Production" not in wbIn.worksheets:
        # only create the sheet if it is not present
        wbIn.create_sheet("Historical_Production")
    ws_historical_prod = wbIn.worksheets[5]  # the sheet that will contain Historical Productions
    # clear the sheet if it is not newly created before copying contents
    for row in ws_historical_prod.iter_rows():
        for cell in row:
            cell.value = None

    # to copy the contents into the main workbook tab, open the source from open refine outputs
    wb_historical_prod_source = openpyxl.load_workbook("./OpenRefine Outputs/Historical Production.xlsx")
    ws_historical_prod_source = wb_historical_prod_source.active  # source from which to be copied

    for row in ws_historical_prod_source.iter_rows(values_only=True):
        ws_historical_prod.append(row)

    # correct formatting
    for i in range(1, ws_historical_prod.max_row + 1):
        for j in range(1, ws_historical_prod.max_column + 1):
            copy_cell_styles(ws_historical_prod_source.cell(i, j), ws_historical_prod.cell(i, j))

    # fix top row for btu tab
    ws_historical_prod.freeze_panes = 'A2'

    # bold the headings
    ws_historical_prod['A1'].font = ws_historical_prod['B1'].font = ws_historical_prod['C1'].font = Font(bold=True)
    ws_historical_prod['D1'].font = ws_historical_prod['E1'].font = ws_historical_prod['F1'].font = ws_historical_prod['G1'].font = Font(bold=True)

    # fix date format
    date_column = 'C'
    date_column_index = column_index_from_string(date_column)

    for row in ws_historical_prod.iter_rows(min_col=date_column_index, max_col=date_column_index, min_row=2, max_row=ws_historical_prod.max_row):
        for cell in row:
            # read the current cell value
            if isinstance(cell.value, str):
                try:
                    # parse the date string and convert to the desired format
                    date = datetime.strptime(cell.value, '%b-%y')
                    # Set the cell value to the new date format (1st day of the month)
                    cell.value = datetime(date.year, date.month, 1)
                    # apply the desired number format to the cell
                    cell.number_format = 'mmm-yy'
                except ValueError:
                    print(f"Skipping cell{cell.coordinate}with value{cell.value}, as it is not a valid date.")

    wb_historical_prod_source.close()  # close the source workbook as it is no longer needed

def populate_btu(wsIn):
    number_line_items = get_number_line_items()
    i = number_line_items + 8
    while i <= wsIn.max_row:
        wsIn.cell(i, 5).value = f'=IFERROR(VLOOKUP($A{i},BTU!$B:$C,2,0),"")'
        wsIn.cell(i, 5).number_format = '#,##0.000_);[Red]\(#,##0.000\)'
        i += number_line_items + 41 # moving to next btu

def populate_first_part(wsIn):
    number_line_items = get_number_line_items()
    i = 4 + number_line_items + 6
    while i <= wsIn.max_row:
        min = i - 6 - number_line_items + 1
        max = i - 6
        for j in range(5, 17): # months
            letter = get_column_letter(j)
            wsIn.cell(i, j).value = f'=IFERROR((SUMIF($D{min}:$D{max},"Oil Sales Revenue ($)",{letter}{min}:{letter}{max})-ABS(SUMIF($D{min}:$D{max},"Oil Revenue Deductions ($)",{letter}{min}:{letter}{max})))/SUMIF($D{min}:$D{max},"Oil Sales Volumes (bbl)",{letter}{min}:{letter}{max}),"")'
            wsIn.cell(i, j).number_format = '#,##0.00_);[Red](#,##0.00)'
        i += number_line_items + 41 # moving to next part

    i = 4 + number_line_items + 7
    while i <= wsIn.max_row:
        min = i - 7 - number_line_items + 1
        max = i - 7
        btu = i - 3
        for j in range(5, 17):  # months
            letter = get_column_letter(j)
            wsIn.cell(i, j).value = f'=IFERROR(((SUMIF($D{min}:$D{max},"Gas Sales Revenue ($)",{letter}{min}:{letter}{max})-ABS(SUMIF($D{min}:$D{max},"Gas Revenue Deductions ($)",{letter}{min}:{letter}{max})))/SUMIF($D{min}:$D{max},"Gas Sales Volumes (mcf)",{letter}{min}:{letter}{max}))/$E{btu},"")'
            wsIn.cell(i, j).number_format = '#,##0.00_);[Red](#,##0.00)'
        i += number_line_items + 41  # next part comes after every 85 rows

    i = 4 + number_line_items + 8
    while i <= wsIn.max_row:
        min = i - 8 - number_line_items + 1
        max = i - 8
        for j in range(5, 17): # months
            letter = get_column_letter(j)
            wsIn.cell(i, j).value = f'=IFERROR((SUMIF($D{min}:$D{max},"NGL Sales Revenue ($)",{letter}{min}:{letter}{max})-ABS(SUMIF($D{min}:$D{max},"NGL Revenue Deductions ($)",{letter}{min}:{letter}{max})))/(SUMIF($D{min}:$D{max},"NGL Sales Volumes (bbl)",{letter}{min}:{letter}{max})+(SUMIF($D{min}:$D{max},"NGL Sales Volumes (gal)",{letter}{min}:{letter}{max})/42)),"")'
            wsIn.cell(i, j).number_format = '#,##0.00_);[Red](#,##0.00)'
        i += number_line_items + 41  # next part comes after every 85 rows

def populate_second_part(wsIn):
    number_line_items = get_number_line_items()
    i = 4 + number_line_items + 10
    while i <= wsIn.max_row:
        value = i - 4
        for j in range(5, 17): # months
            letter = get_column_letter(j)
            wsIn.cell(i, j).value = f'=IFERROR({letter}{value}-{letter}$1,"")'
            wsIn.cell(i, j).number_format = '#,##0.00_);[Red](#,##0.00)'
        num = 14
        for j in range(17, 21): # averages
            letter = get_column_letter(num)
            wsIn.cell(i, j).value = f'=IFERROR(AVERAGE({letter}{i}:P{i}),"")'
            wsIn.cell(i, j).number_format = '#,##0.00_);[Red](#,##0.00)'
            num -= 3
        i += number_line_items + 41 # next part comes after every 85 rows

    i = 4 + number_line_items + 11
    while i <= wsIn.max_row:
        value = i - 4
        for j in range(5, 17): # months
            letter = get_column_letter(j)
            wsIn.cell(i, j).value = f'=IFERROR({letter}{value}-{letter}$2,"")'
            wsIn.cell(i, j).number_format = '#,##0.00_);[Red](#,##0.00)'
        num = 14
        for j in range(17, 21):  # averages
            letter = get_column_letter(num)
            wsIn.cell(i, j).value = f'=IFERROR(AVERAGE({letter}{i}:P{i}),"")'
            wsIn.cell(i, j).number_format = '#,##0.00_);[Red](#,##0.00)'
            num -= 3
        i += number_line_items + 41

    i = 4 + number_line_items + 12
    while i <= wsIn.max_row:
        value = i - 4
        for j in range(5, 17):  # months
            letter = get_column_letter(j)
            wsIn.cell(i, j).value = f'=IFERROR({letter}{value}-{letter}$1,"")'
            wsIn.cell(i, j).number_format = '#,##0.00_);[Red](#,##0.00)'
        num = 14
        for j in range(17, 21):  # averages
            letter = get_column_letter(num)
            wsIn.cell(i, j).value = f'=IFERROR(AVERAGE({letter}{i}:P{i}),"")'
            wsIn.cell(i, j).number_format = '#,##0.00_);[Red](#,##0.00)'
            num -= 3
        i += number_line_items + 41

def populate_third_part(wsIn):
    number_line_items = get_number_line_items()
    i = 4 + number_line_items + 14
    while i <= wsIn.max_row:
        value = i - 8
        for j in range(5, 17): # months
            letter = get_column_letter(j)
            wsIn.cell(i, j).value = f'=IFERROR({letter}{value}/{letter}$1,"")'
            wsIn.cell(i, j).number_format = '0.00%'
        num = 14
        for j in range(17, 21):  # averages
            letter = get_column_letter(num)
            wsIn.cell(i, j).value = f'=IFERROR(AVERAGE({letter}{i}:P{i}),"")'
            wsIn.cell(i, j).number_format = '0.00%'
            num -= 3
        i += number_line_items + 41

    i = 4 + number_line_items + 15
    while i <= wsIn.max_row:
        value = i - 8
        for j in range(5, 17):
            letter = get_column_letter(j)
            wsIn.cell(i, j).value = f'=IFERROR({letter}{value}/{letter}$2,"")'
            wsIn.cell(i, j).number_format = '0.00%'
        num = 14
        for j in range(17, 21):  # averages
            letter = get_column_letter(num)
            wsIn.cell(i, j).value = f'=IFERROR(AVERAGE({letter}{i}:P{i}),"")'
            wsIn.cell(i, j).number_format = '0.00%'
            num -= 3
        i += number_line_items + 41

    i = 4 + number_line_items + 16
    while i <= wsIn.max_row:
        value = i - 8
        for j in range(5, 17):
            letter = get_column_letter(j)
            wsIn.cell(i, j).value = f'=IFERROR({letter}{value}/{letter}$1,"")'
            wsIn.cell(i, j).number_format = '0.00%'
        num = 14
        for j in range(17, 21):  # averages
            letter = get_column_letter(num)
            wsIn.cell(i, j).value = f'=IFERROR(AVERAGE({letter}{i}:P{i}),"")'
            wsIn.cell(i, j).number_format = '0.00%'
            num -= 3
        i += number_line_items + 41

def populate_fourth_part(wsIn):
    number_line_items = get_number_line_items()
    i = 4 + number_line_items + 18
    while i <= wsIn.max_row:
        for j in range(5, 17):
            letter = get_column_letter(j)
            wsIn.cell(i, j).value = f'=_xlfn.XLOOKUP($A{i}&"|"&{letter}$4,Historical_Production!$B:$B&"|"&Historical_Production!$C:$C,Historical_Production!$D:$D,"",0,1)'
            wsIn.cell(i, j).number_format = '#,##0_);[Red](#,##0)'
        i += number_line_items + 41

    i = 4 + number_line_items + 19
    while i <= wsIn.max_row:
        min = i - 19 - number_line_items + 1
        max = i - 19
        value = i - 1
        for j in range(5, 17):
            letter = get_column_letter(j)
            wsIn.cell(i, j).value = f'=IFERROR(SUMIF($D{min}:$D{max},"Gas Sales Volumes (mcf)",{letter}{min}:{letter}{max})/{letter}{value},"")'
            wsIn.cell(i, j).number_format = '0.00%'
        num = 14
        for j in range(17, 21):  # averages
            letter = get_column_letter(num)
            wsIn.cell(i, j).value = f'=IFERROR(AVERAGE({letter}{i}:P{i}),"")'
            wsIn.cell(i, j).number_format = '0.00%'
            num -= 3
        i += number_line_items + 41

def populate_fifth_part(wsIn):
    number_line_items = get_number_line_items()
    i = 4 + number_line_items + 21
    while i <= wsIn.max_row:
        min = i - 21 - number_line_items + 1
        max = i - 21
        value = i - 3
        for j in range(5, 17):
            letter = get_column_letter(j)
            wsIn.cell(i, j).value = f'=IFERROR((SUMIF($D{min}:$D{max},"NGL Sales Volumes (bbl)",{letter}{min}:{letter}{max})+(SUMIF($D{min}:$D{max},"NGL Sales Volumes (gal)",{letter}{min}:{letter}{max})/42))/({letter}{value}/1000),"")'
            wsIn.cell(i, j).number_format = '#,##0.0_);[Red]\(#,##0.0\)'
        num = 14
        for j in range(17, 21):  # averages
            letter = get_column_letter(num)
            wsIn.cell(i, j).value = f'=IFERROR(AVERAGE({letter}{i}:P{i}),"")'
            wsIn.cell(i, j).number_format = '#,##0.0_);[Red]\(#,##0.0\)'
            num -= 3
        i += number_line_items + 41

    i = 4 + number_line_items + 22
    while i <= wsIn.max_row:
        min = i - 22 - number_line_items + 1
        max = i - 22
        value = i - 4
        for j in range(5, 17):
            letter = get_column_letter(j)
            wsIn.cell(i, j).value = f'=IFERROR((SUMIF($D{min}:$D{max},"NGL Sales Volumes (bbl)",{letter}{min}:{letter}{max})+(SUMIF($D{min}:$D{max},"NGL Sales Volumes (gal)",{letter}{min}:{letter}{max})/42))/({letter}{value}),"")'
            wsIn.cell(i, j).number_format = '#,##0.0000_);[Red]\(#,##0.0000\)'
        num = 14
        for j in range(17, 21):  # averages
            letter = get_column_letter(num)
            wsIn.cell(i, j).value = f'=IFERROR(AVERAGE({letter}{i}:P{i}),"")'
            wsIn.cell(i, j).number_format = '#,##0.0000_);[Red]\(#,##0.0000\)'
            num -= 3
        i += number_line_items + 41

def populate_sixth_part(wsIn):
    number_line_items = get_number_line_items()
    i = 4 + number_line_items + 24
    while i <= wsIn.max_row:
        min = i - 24 - number_line_items + 1
        max = i - 24
        for j in range(5, 17):
            letter = get_column_letter(j)
            wsIn.cell(i, j).value = f'=IFERROR(SUMIF($D{min}:$D{max},"Fixed Expense ($)",{letter}{min}:{letter}{max})+SUMIF($D{min}:$D{max},"Oil Variable Expense ($)",{letter}{min}:{letter}{max})+SUMIF($D{min}:$D{max},"Gas Variable Expense ($)",{letter}{min}:{letter}{max}),"")'
            wsIn.cell(i, j).number_format = '#,##0_);[Red](#,##0)'
        i += number_line_items + 41

def populate_seventh_part(wsIn):
    number_line_items = get_number_line_items()
    number = 4 + number_line_items + 26
    for i in range(number, number + 3):
        j = i
        while j <= wsIn.max_row:
            wsIn.cell(j, 5).value = f'=IF(LOS_Designation!$E$5="","",IF(LOS_Designation!$E$5=0,"",VLOOKUP($D{j},LOS_Designation!$D$1:$E$5,2,0)))'
            wsIn.cell(j, 5).number_format = '0%'
            j += number_line_items + 41

def populate_eighth_part(wsIn):
    number_line_items = get_number_line_items()
    i = 4 + number_line_items + 30
    while i <= wsIn.max_row:
        min = i - 30 - number_line_items + 1
        max = i - 30
        value = i - 4
        value_1 = i - 6
        for j in range(5, 17):
            letter = get_column_letter(j)
            wsIn.cell(i, j).value = f'=IF($E{value}<>"",$E{value}*{letter}{value_1},IFERROR(SUMIF($D{min}:$D{max},"Fixed Expense ($)",{letter}{min}:{letter}{max}),""))'
            wsIn.cell(i, j).number_format = '#,##0_);[Red](#,##0)'
        i += number_line_items + 41

    i = 4 + number_line_items + 31
    while i <= wsIn.max_row:
        for j in range(5, 17):
            letter = get_column_letter(j)
            wsIn.cell(i, j).value = f'=_xlfn.XLOOKUP($A{i}&"|"&{letter}$4,Historical_Production!$B:$B&"|"&Historical_Production!$C:$C,Historical_Production!$G:$G,"",0,1)'
            wsIn.cell(i, j).number_format = '#,##0_);[Red](#,##0)'
        i += number_line_items + 41

    i = 4 + number_line_items + 32
    while i <= wsIn.max_row:
        min = i - 2
        max = i - 1
        for j in range(5, 17):
            letter = get_column_letter(j)
            wsIn.cell(i, j).value = f'=IFERROR({letter}{min}/{letter}{max},"")'
            wsIn.cell(i, j).number_format = '#,##0_);[Red](#,##0)'
        num = 14
        for j in range(17, 21):  # averages
            letter = get_column_letter(num)
            wsIn.cell(i, j).value = f'=IFERROR(AVERAGE({letter}{i}:P{i}),"")'
            wsIn.cell(i, j).number_format = '#,##0_);[Red](#,##0)'
            num -= 3
        i += number_line_items + 41

def populate_ninth_part(wsIn):
    number_line_items = get_number_line_items()
    i = 4 + number_line_items + 34
    while i <= wsIn.max_row:
        min = i - 34 - number_line_items + 1
        max = i - 34
        value = i - 7
        value_1 = i - 10
        for j in range(5, 17):
            letter = get_column_letter(j)
            wsIn.cell(i, j).value = f'=IFERROR(IF($E{value}<>"",$E{value}*{letter}{value_1},SUMIF($D{min}:$D{max},"Oil Variable Expense ($)",{letter}{min}:{letter}{max})),"")'
            wsIn.cell(i, j).number_format = '#,##0_);[Red](#,##0)'
        i += number_line_items + 41

    i = 4 + number_line_items + 35
    while i <= wsIn.max_row:
        min = i - 35 - number_line_items + 1
        max = i - 35
        for j in range(5, 17):
            letter = get_column_letter(j)
            wsIn.cell(i, j).value = f'=IFERROR(SUMIF($D{min}:$D{max},"Oil Sales Volumes (bbl)",{letter}{min}:{letter}{max}),"")'
            wsIn.cell(i, j).number_format = '#,##0_);[Red](#,##0)'
        i += number_line_items + 41

    i = 4 + number_line_items + 36
    while i <= wsIn.max_row:
        min = i - 2
        max = i - 1
        for j in range(5, 17):
            letter = get_column_letter(j)
            wsIn.cell(i, j).value = f'=IFERROR({letter}{min}/{letter}{max},"")'
            wsIn.cell(i, j).number_format = '#,##0.00_);[Red](#,##0.00)'
        num = 14
        for j in range(17, 21):  # averages
            letter = get_column_letter(num)
            wsIn.cell(i, j).value = f'=IFERROR(SUM({letter}{min}:P{min})/SUM({letter}{max}:P{max}),"")'
            wsIn.cell(i, j).number_format = '#,##0.00_);[Red](#,##0.00)'
            num -= 3
        i += number_line_items + 41

def populate_last_part(wsIn):
    number_line_items = get_number_line_items()
    i = 4 + number_line_items + 38
    while i <= wsIn.max_row:
        min = i - 38 - number_line_items + 1
        max = i - 38
        value = i - 10
        value_1 = i - 14
        for j in range(5, 17):
            letter = get_column_letter(j)
            wsIn.cell(i, j).value = f'=IFERROR(IF($E{value}<>"",$E{value}*{letter}{value_1},SUMIF($D{min}:$D{max},"Gas Variable Expense ($)",{letter}{min}:{letter}{max})),"")'
            wsIn.cell(i, j).number_format = '#,##0_);[Red](#,##0)'
        i += number_line_items + 41

    i = 4 + number_line_items + 39
    while i <= wsIn.max_row:
        min = i - 39 - number_line_items + 1
        max = i - 39
        for j in range(5, 17):
            letter = get_column_letter(j)
            wsIn.cell(i, j).value = f'=IFERROR(SUMIF($D{min}:$D{max},"Gas Sales Volumes (mcf)",{letter}{min}:{letter}{max}),"")'
            wsIn.cell(i, j).number_format = '#,##0_);[Red](#,##0)'
        i += number_line_items + 41

    i = 4 + number_line_items + 40
    while i <= wsIn.max_row:
        min = i - 2
        max = i - 1
        for j in range(5, 17):
            letter = get_column_letter(j)
            wsIn.cell(i, j).value = f'=IFERROR({letter}{min}/{letter}{max},"")'
            wsIn.cell(i, j).number_format = '#,##0.00_);[Red](#,##0.00)'
        num = 14
        for j in range(17, 21):  # averages
            letter = get_column_letter(num)
            wsIn.cell(i, j).value = f'=IFERROR(SUM({letter}{min}:P{min})/SUM({letter}{max}:P{max}),"")'
            wsIn.cell(i, j).number_format = '#,##0.00_);[Red](#,##0.00)'
            num -= 3
        i += number_line_items + 41

def step_5():
    wbIn = openpyxl.load_workbook("./bot_outputs/step_4_out.xlsx")
    add_btu_tab(wbIn)
    add_historical_prod_tab(wbIn)

    wsIn = wbIn.worksheets[0] # active LOS worksheet

    # populate values
    populate_btu(wsIn)
    populate_first_part(wsIn)
    populate_second_part(wsIn)
    populate_third_part(wsIn)
    populate_fourth_part(wsIn)
    populate_fifth_part(wsIn)
    populate_sixth_part(wsIn)
    populate_seventh_part(wsIn)
    populate_eighth_part(wsIn)
    populate_ninth_part(wsIn)
    populate_last_part(wsIn)

    # now add filters
    wsIn.auto_filter.ref = 'A4:P4'

    wbIn.save("./bot_outputs/step_5_out.xlsx")

step_5()