import openpyxl
from openpyxl.styles import Font
from step_0 import copy_cell_styles
from openpyxl.utils import get_column_letter

def add_pricing_inputs_tab(wbIn):
    if "Historical_NYMEX_Pricing_Input" not in wbIn.worksheets:
        # only create the sheet if it is not present
        wbIn.create_sheet("Historical_NYMEX_Pricing_Input")
    ws_pricing_input = wbIn.worksheets[3] # the sheet that will contain Pricing Inputs
    # clear the sheet if it is not newly created before copying contents
    for row in ws_pricing_input.iter_rows():
        for cell in row:
            cell.value = None

    # to copy the contents into the main workbook tab, open the source from open refine outputs
    wb_pricing_input_source = openpyxl.load_workbook("./example_los/OpenRefine Outputs/Historical NYMEX Pricing Input.xlsx")
    ws_pricing_input_source = wb_pricing_input_source.active # source from which to be copied

    for row in ws_pricing_input_source.iter_rows(values_only=True):
        ws_pricing_input.append(row)
    
    # correct formatting
    for i in range(1, ws_pricing_input.max_row + 1):
        for j in range(1, ws_pricing_input.max_column + 1):
            copy_cell_styles(ws_pricing_input_source.cell(i, j), ws_pricing_input.cell(i, j))

    # fix top row for pricing inputs tab
    ws_pricing_input.freeze_panes = 'A2'

    # bold the headings
    ws_pricing_input['A1'].font = ws_pricing_input['B1'].font = ws_pricing_input['C1'].font = Font(bold=True)

    wb_pricing_input_source.close() # close the source workbook as it is no longer needed

def step_3():
    wbIn = openpyxl.load_workbook("./bot_outputs/step_2_out.xlsx")
    add_pricing_inputs_tab(wbIn)

    wsIn = wbIn.worksheets[0] # active LOS worksheet

    wsIn.insert_rows(1, 3)
    wsIn.auto_filter.ref = 'A4:P4'
    wsIn.freeze_panes = 'E5'


    wsIn['D1'].value = 'NYMEX Oil'
    wsIn['D1'].font = Font(bold=True)
    wsIn['D2'].value = 'NYMEX Gas'
    wsIn['D2'].font = Font(bold=True)

    # nymex oil and gas values
    for i in range(5, wsIn.max_column + 1):
        wsIn.cell(1, i).value = f'=VLOOKUP({get_column_letter(i)}$4,Historical_NYMEX_Pricing_Input!$A:$C,2,0)'
        wsIn.cell(1, i).number_format = '0.00'
    for i in range(5, wsIn.max_column + 1):
        wsIn.cell(2, i).value = f'=VLOOKUP({get_column_letter(i)}$4,Historical_NYMEX_Pricing_Input!$A:$C,3,0)'
        wsIn.cell(2, i).number_format = '0.00'


    wbIn.save("./bot_outputs/step_3_out.xlsx")

step_3()