import openpyxl
from openpyxl.styles import Font
from step_0 import get_number_line_items

def fix_vlookups(wsIn):
    number_line_items = get_number_line_items()

    count = 0
    i = 5
    while i <= wsIn.max_row:
        if count < number_line_items:
            wsIn.cell(i, 1).value = f'=VLOOKUP($B{i},NameIDRecon!$B:$C,2,0)'
            wsIn.cell(i, 4).value = f'=IF(VLOOKUP($C{i},LOS_Designation!$A:$B,2,0)=0,"",VLOOKUP($C{i},LOS_Designation!$A:$B,2,0))'
            count += 1
            i += 1
        else:
            count = 0
            i += 41

def step_4():
    wbIn = openpyxl.load_workbook("./bot_outputs/step_3_out.xlsx")
    
    wsIn = wbIn.worksheets[0]

    columns = list(wsIn.iter_cols(min_col=17, max_col=20, min_row=4, max_row=4))
    for column, month in zip(columns, [3, 6, 9, 12]):
        column[0].value = f"{month}-Mo Avg"
        column[0].font = Font(bold=True)

    items = [
        None,
        "Vinci",
        None,
        "BTU",
        None,
        "Oil Price ($/bbl)",
        "Gas Price ($/mmbtu)",
        "NGL Price ($/bbl)",
        None,
        "Oil Differential ($/bbl)",
        "Gas Differential ($/mmbtu)",
        "NGL Differential ($/bbl)",
        None,
        "Oil Differential (%)",
        "Gas Differential (%)",
        "NGL Differential (%)",
        None,
        "Gross Historical Gas Production (mcf)",
        "Shrink (% remaining)",
        None,
        "NGL Yield (bbl/mmcf)",
        "NGL Yield (bbl/mcf)",
        None,
        "Total Expenses ($/mo)",
        None,
        "Fixed Expense (%)",
        "Oil Variable Expense (%)",
        "Gas Variable Expense (%)",
        None,
        "Fixed Expenses ($/mo)",
        "Well Count",
        "Fixed Expense ($/well/mo)",
        None,
        "Oil Variable Expenses ($/mo)",
        "Gross Oil Sales Volumes (bbl)",
        "Oil Variable Expense ($/bbl)",
        None,
        "Gas Variable Expenses ($/mo)",
        "Gross Gas Sales Volumes (mcf)",
        "Gas Variable Expense ($/mcf)",
        None
    ]
    # 41 is the number of rows which will now be inserted below every record
    # inserting blank rows first, then populate them
    inserting_rows = 41
    number_line_items = get_number_line_items()
    count = number_line_items
    for i in range(wsIn.max_row, 4, -1):
        if count == number_line_items:
            count = 1
            wsIn.insert_rows(i + 1, inserting_rows)
            # we will populate the rows at the time of insertion
            row_start = i + 1
            for item in items:
                wsIn.cell(row_start, 4).value = item
                if items.index(item) == 1:
                    wsIn.cell(row_start, 4).font = Font(bold=True)
                row_start += 1

        else:
            count += 1

    # add names and id's in columns
    for i in range(5, wsIn.max_row + 1):
        # find the first empty cell, then start populating the whole column with the formula
        if wsIn.cell(i, 1).value == None:
            value = i - 1
            formula_A = f'=A{value}'
            formula_B = f'=B{value}'
            count = 0 # to prevent the while loop going infinite
            while wsIn.cell(i, 1).value == None and count < 41:
                wsIn.cell(i, 1).value = formula_A
                wsIn.cell(i, 2).value = formula_B
                i += 1
                count += 1

    # remove empty row inserted at last
    if all(cell.value is None for cell in wsIn[wsIn.max_row]):
        wsIn.delete_rows(wsIn.max_row, 1)

    # fix vlookups
    fix_vlookups(wsIn)

    # adjust_column_widths(wsIn)
    # print(wsIn.max_row)
    wbIn.save("./bot_outputs/step_4_out.xlsx")

step_4()