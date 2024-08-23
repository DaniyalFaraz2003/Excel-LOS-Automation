import openpyxl
from openpyxl import Workbook
from copy import copy

def copy_cell_styles(source_cell, target_cell):
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)


def get_rows_of_name(sheet, search_value):
    matching_rows = []
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        if row[0].value == search_value:
            matching_rows.append([cell.value for cell in row])
    
    return matching_rows

def store_number_line_items(ALL_LINE_ITEMS):
    number_line_items = len(ALL_LINE_ITEMS) + sum(len(v) for v in ALL_LINE_ITEMS.values())
    wb_number_line_items = Workbook()
    ws_number_line_items = wb_number_line_items.active
    ws_number_line_items['A1'].value = number_line_items
    wb_number_line_items.save('./bot_outputs/number_line_items.xlsx')
    wb_number_line_items.close()

def get_number_line_items():
    wb = openpyxl.load_workbook('./bot_outputs/number_line_items.xlsx')
    ws = wb.active
    number_line_items = ws['A1'].value
    wb.close()
    return number_line_items

def get_all_line_items():
    ALL_LINE_ITEMS = {}
    wb = openpyxl.load_workbook('./OpenRefine Outputs/LOS Designation.xlsx')
    ws = wb.active
    i = 2
    while i <= ws.max_row:
        if ws.cell(i, 1).value[-1] == ':' and ws.cell(i, 2).value == None:
            category = ws.cell(i, 1).value
            ALL_LINE_ITEMS[category] = []
            i += 1
            while ws.cell(i, 1).value[-1] != ':':
                ALL_LINE_ITEMS[category].append(ws.cell(i, 1).value)
                i += 1
                if i > ws.max_row:
                    break
    return ALL_LINE_ITEMS

def step_0():
    wbOrg = openpyxl.load_workbook("./OpenRefine Outputs/LOS.xlsx")
    wbOut = Workbook()
    wsOut = wbOut.active
    wsOrg = wbOrg.active  # active worksheet
    # extract all names from first column to find total number of records
    first_column = [cell.value for cell in wsOrg['A'][1:]]
    names = list({k: None for k in first_column}.keys())

    ALL_LINE_ITEMS = get_all_line_items()

    # calculate number of records and store it in a file
    store_number_line_items(ALL_LINE_ITEMS)
    
    for c in range(1, wsOrg.max_column + 1):
        wsOut.cell(1, c).value = wsOrg.cell(1, c).value
        copy_cell_styles(wsOrg.cell(1, c), wsOut.cell(1, c))
        
    row = 2
    for name in names:
        name_rows = get_rows_of_name(wsOrg, name)
        for category_heading in ALL_LINE_ITEMS.keys():
            wsOut.cell(row, 1).value, wsOut.cell(row, 2).value = name, category_heading
            row += 1
            for category in ALL_LINE_ITEMS[category_heading]:
                wsOut.cell(row, 1).value, wsOut.cell(row, 2).value = name, category
                
                found = False
                for _row in name_rows:
                    if _row[1] == category:
                        found = True
                        for c in range(3, wsOrg.max_column + 1):
                            wsOut.cell(row, c).value = _row[c - 1]
                            wsOut.cell(row, c).number_format = '#,##0.00_);[Red](#,##0.00)'
                        break

                if not found:
                    for c in range(3, wsOrg.max_column + 1):
                        wsOut.cell(row, c).value = 0.00
                        wsOut.cell(row, c).number_format = '#,##0.00_);[Red](#,##0.00)'
                row += 1


    # fix the top row
    wsOut.freeze_panes = 'A2'

    # rename sheet name
    wsOut.title = 'LOS'

    wbOut.save("./bot_outputs/step_0_out.xlsx")


step_0()