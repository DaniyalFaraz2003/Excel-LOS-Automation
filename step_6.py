import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string

def add_economic_param_tab(wbIn):
    if "Economic Parameters" not in wbIn.worksheets:
        # only create the sheet if it is not present
        wbIn.create_sheet("Economic Parameters", 0) # first position
    ws_economic_param = wbIn.worksheets[0]  # the sheet that will contain Economic Parameters
    # clear the sheet if it is not newly created before copying contents
    for row in ws_economic_param.iter_rows():
        for cell in row:
            cell.value = None

def get_names_ids():
    wb_names_ids = openpyxl.load_workbook("./example_los/OpenRefine Outputs/Example0gross_NameIDReconciliation.xlsx")
    ws_names_ids = wb_names_ids.active
    names = [cell.value for cell in ws_names_ids['B'][1:]]
    ids = [cell.value for cell in ws_names_ids['C'][1:]]
    return names, ids

def insert_headings(wsIn):
    wsIn.merge_cells('A1:B1')
    wsIn['A2'].value = 'PHDWIN Id'
    wsIn['A2'].font = Font(bold=True)
    wsIn['B2'].value = 'LOS Name'
    wsIn['B2'].font = Font(bold=True)
    headings = [
        "Oil Differential ($/bbl)",
        "Gas Differential ($/mmbtu)",
        "NGL Differential ($/bbl)",
        "Oil Differential (%)",
        "Gas Differential (%)",
        "NGL Differential (%)",
        "Shrink (% remaining)",
        "NGL Yield (bbl/mmcf)",
        "NGL Yield (bbl/mcf)",
        "Fixed Expense ($/well/mo)",
        "Oil Variable Expense ($/bbl)",
        "Gas Variable Expense ($/mcf)"
    ]

    i = 3
    for heading in headings:
        letter = get_column_letter(i)
        letter_1 = get_column_letter(i + 3)
        wsIn[f'{letter}1'].value = heading
        wsIn[f'{letter}1'].font = Font(bold=True)
        wsIn[f'{letter}1'].alignment = Alignment(horizontal='center')
        wsIn.merge_cells(f'{letter}1:{letter_1}1')
        i += 4

    wsIn.freeze_panes = 'A2'

def insert_average_headings(wsIn):
    thick_border = Border(right=Side(style='thick'))
    for x in range(1, wsIn.max_row + 1):
        wsIn.cell(x, 2).border = thick_border

    averages_headings = ["3-Mo Avg", "6-Mo Avg", "9-Mo Avg", "12-Mo Avg"]
    i = 0
    for x in range(3, wsIn.max_column + 1):
        letter = get_column_letter(x)
        wsIn[f'{letter}2'] = averages_headings[i]
        wsIn[f'{letter}2'].font = Font(bold=True)
        if i == 3:
            # apply border
            for j in range(1, wsIn.max_row + 1):
                wsIn[f'{letter}{j}'].border = thick_border
        i += 1
        if i > 3:
            i = 0

def fill_names_ids(wsIn):
    names, ids = get_names_ids() # names and ids
    for name, id in zip(names, ids):
        wsIn.append([id, name])

def fill_gray_color(wsIn):
    gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    i = 50
    while i <= wsIn.max_row:
        fill_range = wsIn[f'D{i}':f'T{i + 39}']
        for row in fill_range:
            for cell in row:
                cell.fill = gray_fill
        i += 85

def step_6():
    wbIn = openpyxl.load_workbook("./bot_outputs/step_5_out.xlsx")
    add_economic_param_tab(wbIn)

    wsIn = wbIn.worksheets[0] # economic worksheet

    # insert headings and fill names and ids
    insert_headings(wsIn)
    fill_names_ids(wsIn)
    insert_average_headings(wsIn)

    # now start populating values
    i = 3
    skipper = 0
    while i < wsIn.max_column:
        section_letter = get_column_letter(i)
        for j in range(3, wsIn.max_row + 1):
            value = 'Q'
            for k in range(3, 7):
                wsIn.cell(j, skipper + k).value = f"=_xlfn.XLOOKUP('Economic Parameters'!$A{j} & 'Economic Parameters'!${section_letter}$1, Example0gross_LOS!$A:$A & Example0gross_LOS!$D:$D, Example0gross_LOS!${value}:${value}, "", 0, 1)"
                value = get_column_letter(column_index_from_string(value) + 1) # update next letter
                # now check number formatting
                if i < 15 or i >= 43:
                    wsIn.cell(j, skipper + k).number_format = '#,##0.00_);[Red](#,##0.00)'
                elif i >= 15 and i < 31:
                    wsIn.cell(j, skipper + k).number_format = '0.00%'
                elif i == 31:
                    wsIn.cell(j, skipper + k).number_format = '#,##0.0_);[Red]\(#,##0.0\)'
                elif i == 35:
                    wsIn.cell(j, skipper + k).number_format = '#,##0.0000_);[Red]\(#,##0.0000\)'
                elif i == 39:
                    wsIn.cell(j, skipper + k).number_format = '#,##0_);[Red](#,##0)'
        skipper += 4
        i += 4 # incrementing for outermost loop

    wsIn.freeze_panes = 'C2'

    # fill gray color in LOS worksheet
    ws_los = wbIn.worksheets[1]
    fill_gray_color(ws_los)

    wbIn.active = wbIn.worksheets[0]
    wbIn.save("./bot_outputs/step_6_out.xlsx")

step_6()