import openpyxl
from openpyxl import load_workbook
import logging

def log_details():
    logging.basicConfig(
        filename="./bot_outputs/errors.log",
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S %p',
        filemode='w'
    )
    return logging.getLogger()

def check_empty_rows(rows, logger, file_name):
    for row in rows:
        empty_row = True
        for cell in row:
            if cell.value is not None:
                empty_row = False
                break
        if empty_row == True:
            row_number = rows.index(row) + 1
            logger.info(f'Empty row number at {row_number} in {file_name} File. No empty rows should be present in the data')

def check_empty_cells(data_rows, logger, file_name):
    for row in data_rows:
        for cell in row:
            if cell.value is None:
                logger.info(f'Empty cell(s) present in row {data_rows.index(row) + 2} in {file_name} File')
                break

def validate_los(ws, logger):
    if ws.max_column != 14:
        logger.info("Columns are not correct, check for extra or missing columns in LOS File")

    if ws['A1'].value != 'Account':
        logger.info('Account Names should be in first column of LOS File, and heading name should be Account')
    if ws['B1'].value != 'Description':
        logger.info('Descriptions should be in second column of LOS File, and heading name should be Description')

    rows = list(ws.rows)
    check_empty_rows(rows, logger, 'LOS')

def validate_name_id(ws, logger):
    if ws.max_column != 3:
        logger.info('Total number of columns in Name Id Recon File should be 3')
    if ws['A1'].value != 'Database Name':
        logger.info('Database Names should be in the first column in Name Id Recon File, and heading name should be Database Name')
    if ws['B1'].value != 'LOS Name':
        logger.info('LOS Names should be in the second column in Name Id Recon File, and heading name should be LOS Name')
    if ws['C1'].value != 'PHDWIN Id':
        logger.info('PHDWIN Ids should be in the third column in Name Id Recon File, and heading name should be PHDWIN Id')
    
    data_rows = list(ws.iter_rows(min_row=2))
    check_empty_cells(data_rows, logger, 'Name ID Reconciliation')

def validate_los_desig(ws, logger):
    if ws.max_column != 3:
        logger.info('Total number of columns in LOS Designation File should be 3')
    if ws['A1'].value != 'Description':
        logger.info('All Descriptions should be in first column in LOS Designation File, and heading name should be Description')
    if ws['B1'].value != 'LOS Designation':
        logger.info('All LOS Designations should be in second column in LOS Designation File, and heading name should be LOS Designation')
    if ws['D1'].value != 'Expense %':
        logger.info('Expense % till Total % headings should be in fourth column in LOS Designation File')

    rows = list(ws.iter_rows(max_col=2))
    check_empty_rows(rows, logger, 'LOS Designation')

def validate_hist_nymex(ws, logger):
    if ws.max_column != 3:
        logger.info('Total number of columns in Historical Nymex Input File should be 3')
    if ws['A1'].value != 'Date':
        logger.info('All Dates should be in first column in Historical NYMEX Input File, and heading name should be Date')
    if 'Oil Price'.upper() not in ws['B1'].value.upper():
        logger.info('All Oil Prices should be in second column in Historical NYMEX Input File')
    if 'Gas Price'.upper() not in ws['C1'].value.upper():
        logger.info('All Gas Prices should be in second column in Historical NYMEX Input File')

    data_rows = list(ws.iter_rows(min_row=2))
    check_empty_cells(data_rows, logger, 'Historical NYMEX Input')

def validate_btu(ws, logger):
    if ws.max_column != 3:
        logger.info('Total number of columns in BTU File should be 3')
    if ws['A1'].value != 'Case Name':
        logger.info('Case Names should be in the first column in BTU File, and heading name should be Case Name')
    if ws['B1'].value != 'PHDWIN Id':
        logger.info('PHDWIN Ids should be in the second column in BTU File, and heading name should be PHDWIN Id')
    if ws['C1'].value != 'BTU Factor':
        logger.info('BTU Factors should be in the third column in BTU File, and heading name should be BTU Factor')
    
    data_rows = list(ws.iter_rows(min_row=2))
    check_empty_cells(data_rows, logger, 'BTU')

def validate_hist_prod(ws, logger):
    if ws.max_column != 7:
        logger.info('Total number of columns in Historical Production File should be 7')
    if ws['A1'].value != 'Case Name':
        logger.info('Case Names should be in the first column in Historical Production File, and heading name should be Case Name')
    if ws['B1'].value != 'PHDWIN Id':
        logger.info('PHDWIN Ids should be in the second column in Historical Production File, and heading name should be PHDWIN Id')
    if ws['C1'].value != 'Date':
        logger.info('Dates should be in the third column in Historical Production File, and heading name should be Date')
    if 'Gas Production'.upper() not in ws['D1'].value.upper():
        logger.info('Gas Productions should be in the fourth column in Historical Production File')
    if 'Oil Production'.upper() not in ws['E1'].value.upper():
        logger.info('Oil Productions should be in the fifth column in Historical Production File')
    if 'Water Production'.upper() not in ws['F1'].value.upper():
        logger.info('Water Productions should be in the sixth column in Historical Production File')
    if 'Well Count'.upper() not in ws['G1'].value.upper():
        logger.info('Well Counts should be in the seventh column in Historical Production File')
    
    data_rows = list(ws.iter_rows(min_row=2))
    check_empty_cells(data_rows, logger, 'Historical Production')

def validate_files():
    logger = log_details()

    wb_los = load_workbook('./OpenRefine Outputs/LOS.xlsx')
    wb_name_id = load_workbook('./OpenRefine Outputs/NameIDReconciliation.xlsx')
    wb_los_desig = load_workbook('./OpenRefine Outputs/LOS Designation.xlsx')
    wb_hist_nymex = load_workbook('./OpenRefine Outputs/Historical NYMEX Input.xlsx')
    wb_btu = load_workbook('./OpenRefine Outputs/BTU.xlsx')
    wb_hist_prod = load_workbook('./OpenRefine Outputs/Historical Production.xlsx')
    
    ws_los = wb_los.active
    ws_name_id = wb_name_id.active
    ws_los_desig = wb_los_desig.active
    ws_hist_nymex = wb_hist_nymex.active
    ws_btu = wb_btu.active
    ws_hist_prod = wb_hist_prod.active

    validate_los(ws_los, logger)
    validate_name_id(ws_name_id, logger)
    validate_los_desig(ws_los_desig, logger)
    validate_hist_nymex(ws_hist_nymex, logger)
    validate_btu(ws_btu, logger)
    validate_hist_prod(ws_hist_prod, logger)



validate_files()